import io
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from models import UserRegistration, Feedback

logger = logging.getLogger(__name__)


def generate_excel_report():
    """Generate Excel file with user registrations and feedback data"""
    try:
        # Create a new workbook
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Create User Registrations sheet
        ws1 = wb.create_sheet("User Registrations")

        # Headers for User Registrations
        headers1 = ['ID', 'Name', 'Email', 'Phone', 'Gender', 'Profession', 'User Type', 'Submitted At', 'IP Address']
        ws1.append(headers1)

        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        for cell in ws1[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Get user registration data
        registrations, _ = UserRegistration.get_all(page=1, per_page=10000)  # Get all records
        for reg in registrations:
            ws1.append([
                reg.id,
                reg.name,
                reg.email,
                reg.phone,
                reg.gender or '',
                reg.profession or '',
                reg.user_type,
                reg.submitted_at.strftime('%Y-%m-%d %H:%M:%S') if reg.submitted_at else '',
                reg.ip_address or ''
            ])

        # Auto-adjust column widths
        for column in ws1.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws1.column_dimensions[column_letter].width = adjusted_width

        # Create Feedback Submissions sheet
        ws2 = wb.create_sheet("Feedback Submissions")

        # Headers for Feedback
        headers2 = [
            'ID', 'Visual Design', 'Visual Design Issue', 'Ease of Navigation', 'Navigation Issue',
            'Mobile Responsiveness', 'Mobile Issue', 'Overall Satisfaction', 'Satisfaction Issue',
            'Ease of Tasks', 'Tasks Issue', 'Quality of Services', 'Services Issue',
            'Like Most', 'Improvements', 'Features', 'Legal Challenges', 'Additional Comments',
            'Contact Willing', 'Contact Email', 'Submitted At', 'IP Address'
        ]
        ws2.append(headers2)

        # Style headers
        for cell in ws2[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Get feedback data
        feedback_list, _ = Feedback.get_all(page=1, per_page=10000)  # Get all records
        for feedback in feedback_list:
            ws2.append([
                feedback.id,
                feedback.visual_design or '',
                feedback.visual_design_issue or '',
                feedback.ease_of_navigation or '',
                feedback.ease_of_navigation_issue or '',
                feedback.mobile_responsiveness or '',
                feedback.mobile_responsiveness_issue or '',
                feedback.overall_satisfaction or '',
                feedback.overall_satisfaction_issue or '',
                feedback.ease_of_tasks or '',
                feedback.ease_of_tasks_issue or '',
                feedback.quality_of_services or '',
                feedback.quality_of_services_issue or '',
                feedback.like_most or '',
                feedback.improvements or '',
                feedback.features or '',
                feedback.legal_challenges or '',
                feedback.additional_comments or '',
                feedback.contact_willing or '',
                feedback.contact_email or '',
                feedback.submitted_at.strftime('%Y-%m-%d %H:%M:%S') if feedback.submitted_at else '',
                feedback.ip_address or ''
            ])

        # Auto-adjust column widths
        for column in ws2.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws2.column_dimensions[column_letter].width = adjusted_width

        # Save to BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        return excel_buffer

    except Exception as e:
        logger.error(f'Error generating Excel report: {str(e)}')
        return None
